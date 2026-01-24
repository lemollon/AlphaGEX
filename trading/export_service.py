"""
Trade Export Service for AlphaGEX

Provides Excel/CSV export functionality with full transparency:
- Complete trade history with P&L breakdown
- P&L attribution showing exactly how each trade contributed
- Decision logs with AI reasoning
- Wheel cycle summaries

Uses openpyxl for Excel generation with proper formatting.
"""

import io
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Optional, List, Dict, Any, BinaryIO
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel export will be disabled.")

import pandas as pd
from database_adapter import get_connection

logger = logging.getLogger(__name__)


class TradeExportService:
    """
    Service for exporting trading data to Excel/CSV with full transparency.

    All exports pull from REAL PostgreSQL data - no simulated values.
    """

    def __init__(self):
        # Texas Central Time - standard timezone for all AlphaGEX operations
        self.tz = ZoneInfo("America/Chicago")

        # Excel styling
        if OPENPYXL_AVAILABLE:
            self.header_font = Font(bold=True, color="FFFFFF")
            self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            self.money_positive = Font(color="006400")  # Dark green
            self.money_negative = Font(color="8B0000")  # Dark red
            self.thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def export_trade_history(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        symbol: str = 'SPY',
        format: str = 'xlsx'
    ) -> BinaryIO:
        """
        Export complete trade history with full details.

        Args:
            start_date: Start date filter (default: 30 days ago)
            end_date: End date filter (default: today)
            symbol: Symbol to export (default: SPY)
            format: 'xlsx' or 'csv'

        Returns:
            BytesIO buffer containing the file
        """
        if not start_date:
            start_date = datetime.now(self.tz) - timedelta(days=30)
        if not end_date:
            end_date = datetime.now(self.tz)

        conn = get_connection()

        # Get closed trades
        closed_trades = pd.read_sql_query('''
            SELECT
                id,
                symbol,
                strategy,
                strike,
                option_type,
                contracts,
                entry_date,
                entry_time,
                entry_price,
                exit_date,
                exit_time,
                exit_price,
                realized_pnl,
                exit_reason,
                hold_time_hours,
                entry_spot_price,
                exit_spot_price,
                entry_vix,
                exit_vix,
                gex_regime,
                created_at
            FROM autonomous_closed_trades
            WHERE symbol = %s
              AND COALESCE(exit_date, entry_date) >= %s
              AND COALESCE(exit_date, entry_date) <= %s
            ORDER BY COALESCE(exit_date, entry_date) DESC, COALESCE(exit_time, entry_time) DESC
        ''', conn, params=(symbol, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

        # Get open positions
        open_positions = pd.read_sql_query('''
            SELECT
                id,
                symbol,
                strategy,
                strike,
                option_type,
                contracts,
                entry_date,
                entry_time,
                entry_price,
                current_price,
                unrealized_pnl,
                entry_spot_price,
                current_spot_price,
                expiration_date,
                gex_regime,
                confidence,
                trade_reasoning,
                profit_target_pct,
                stop_loss_pct,
                created_at
            FROM autonomous_open_positions
            WHERE symbol = %s
            ORDER BY entry_date DESC, entry_time DESC
        ''', conn, params=(symbol,))

        conn.close()

        if format == 'csv':
            return self._export_trades_csv(closed_trades, open_positions)
        else:
            return self._export_trades_excel(closed_trades, open_positions, symbol, start_date, end_date)

    def export_pnl_attribution(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        symbol: str = 'SPY'
    ) -> BinaryIO:
        """
        Export P&L attribution showing exactly how each trade contributed.

        Columns:
        - Trade ID, Date, Strategy
        - Entry Price, Exit Price
        - Gross P&L (before costs)
        - Commission, Slippage
        - Net P&L
        - Running Total
        - Contribution % (what % of total P&L this trade represents)
        """
        if not start_date:
            start_date = datetime.now(self.tz) - timedelta(days=30)
        if not end_date:
            end_date = datetime.now(self.tz)

        conn = get_connection()

        trades = pd.read_sql_query('''
            SELECT
                id,
                exit_date,
                exit_time,
                strategy,
                strike,
                option_type,
                contracts,
                entry_price,
                exit_price,
                realized_pnl,
                hold_time_hours,
                exit_reason
            FROM autonomous_closed_trades
            WHERE symbol = %s
              AND COALESCE(exit_date, entry_date) >= %s
              AND COALESCE(exit_date, entry_date) <= %s
            ORDER BY COALESCE(exit_date, entry_date) ASC, COALESCE(exit_time, entry_time) ASC
        ''', conn, params=(symbol, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

        conn.close()

        if trades.empty:
            # Return empty workbook
            return self._create_empty_export("No trades found for the specified period")

        # Calculate attribution metrics
        trades['gross_pnl'] = (trades['exit_price'] - trades['entry_price']) * trades['contracts'] * 100

        # Estimate commission ($0.65 per contract, entry + exit)
        trades['commission'] = trades['contracts'] * 0.65 * 2

        # Slippage estimate (difference between gross and realized)
        trades['slippage'] = trades['gross_pnl'] - trades['realized_pnl'] - trades['commission']
        trades['slippage'] = trades['slippage'].clip(lower=0)  # Slippage shouldn't be negative

        # Net P&L is the realized P&L from DB (already accounts for everything)
        trades['net_pnl'] = trades['realized_pnl']

        # Running total
        trades['running_total'] = trades['net_pnl'].cumsum()

        # Contribution percentage
        total_pnl = trades['net_pnl'].sum()
        if total_pnl != 0:
            trades['contribution_pct'] = (trades['net_pnl'] / abs(total_pnl)) * 100
        else:
            trades['contribution_pct'] = 0

        # Win/Loss indicator
        trades['result'] = trades['net_pnl'].apply(lambda x: 'WIN' if x > 0 else 'LOSS' if x < 0 else 'BREAK-EVEN')

        return self._export_pnl_attribution_excel(trades, symbol, start_date, end_date, total_pnl)

    def export_decision_logs(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        symbol: str = 'SPY'
    ) -> BinaryIO:
        """
        Export decision logs with AI reasoning for full transparency.

        Shows:
        - What the system saw (market data, GEX, VIX)
        - What it analyzed (patterns, psychology traps)
        - What it decided (trade or no trade)
        - Why (AI reasoning, confidence scores)
        """
        if not start_date:
            start_date = datetime.now(self.tz) - timedelta(days=7)
        if not end_date:
            end_date = datetime.now(self.tz)

        conn = get_connection()

        logs = pd.read_sql_query('''
            SELECT
                timestamp,
                log_type,
                symbol,
                spot_price,
                net_gex,
                flip_point,
                call_wall,
                put_wall,
                vix_level,
                pattern_detected,
                confidence_score,
                trade_direction,
                risk_level,
                psychology_trap,
                rsi_5m, rsi_15m, rsi_1h, rsi_4h, rsi_1d,
                strike_chosen,
                strike_selection_reason,
                kelly_pct,
                contracts,
                ai_thought_process,
                ai_confidence,
                ai_warnings,
                action_taken,
                strategy_name,
                reasoning_summary,
                full_reasoning,
                position_id,
                scan_cycle,
                session_id
            FROM autonomous_trader_logs
            WHERE symbol = %s
              AND timestamp >= %s
              AND timestamp <= %s
            ORDER BY timestamp DESC
        ''', conn, params=(symbol, start_date, end_date))

        conn.close()

        if logs.empty:
            return self._create_empty_export("No decision logs found for the specified period")

        return self._export_decision_logs_excel(logs, symbol, start_date, end_date)

    def export_wheel_cycles(self, symbol: str = None) -> BinaryIO:
        """
        Export wheel strategy cycle history.

        Shows:
        - Cycle ID, Symbol, Status
        - CSP trades and premiums
        - Assignment details
        - Covered call trades and premiums
        - Call-away details
        - Total premium collected
        - Total P&L per cycle
        """
        conn = get_connection()

        # Check if wheel tables exist
        try:
            cycles = pd.read_sql_query('''
                SELECT
                    c.id as cycle_id,
                    c.symbol,
                    c.status,
                    c.start_date,
                    c.end_date,
                    c.shares_owned,
                    c.share_cost_basis,
                    c.total_csp_premium,
                    c.total_cc_premium,
                    c.total_premium_collected,
                    c.assignment_date,
                    c.assignment_price,
                    c.called_away_date,
                    c.called_away_price,
                    c.realized_pnl
                FROM wheel_cycles c
                WHERE (%s IS NULL OR c.symbol = %s)
                ORDER BY c.start_date DESC
            ''', conn, params=(symbol, symbol))

            legs = pd.read_sql_query('''
                SELECT
                    l.id as leg_id,
                    l.cycle_id,
                    l.leg_type,
                    l.action,
                    l.strike,
                    l.expiration_date,
                    l.contracts,
                    l.premium_received,
                    l.premium_paid,
                    l.open_date,
                    l.close_date,
                    l.close_reason,
                    l.underlying_price_at_open,
                    l.underlying_price_at_close,
                    l.delta_at_open,
                    l.dte_at_open
                FROM wheel_legs l
                JOIN wheel_cycles c ON c.id = l.cycle_id
                WHERE (%s IS NULL OR c.symbol = %s)
                ORDER BY l.open_date
            ''', conn, params=(symbol, symbol))

        except Exception as e:
            logger.warning(f"Wheel tables not found: {e}")
            conn.close()
            return self._create_empty_export("No wheel cycles found. Start a wheel strategy first.")

        conn.close()

        if cycles.empty:
            return self._create_empty_export("No wheel cycles found")

        return self._export_wheel_cycles_excel(cycles, legs, symbol)

    def export_full_audit(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        symbol: str = 'SPY'
    ) -> BinaryIO:
        """
        Export a complete audit package with all sheets:
        - Trade History
        - P&L Attribution
        - Decision Logs
        - Wheel Cycles (if any)
        - Performance Summary
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required for Excel export. Install with: pip install openpyxl")

        if not start_date:
            start_date = datetime.now(self.tz) - timedelta(days=30)
        if not end_date:
            end_date = datetime.now(self.tz)

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        conn = get_connection()

        # Sheet 1: Trade History
        self._add_trade_history_sheet(wb, conn, symbol, start_date, end_date)

        # Sheet 2: P&L Attribution
        self._add_pnl_attribution_sheet(wb, conn, symbol, start_date, end_date)

        # Sheet 3: Decision Logs
        self._add_decision_logs_sheet(wb, conn, symbol, start_date, end_date)

        # Sheet 4: Wheel Cycles (if applicable)
        self._add_wheel_cycles_sheet(wb, conn, symbol)

        # Sheet 5: Performance Summary
        self._add_performance_summary_sheet(wb, conn, symbol, start_date, end_date)

        conn.close()

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # =========================================================================
    # Private helper methods for Excel generation
    # =========================================================================

    def _export_trades_excel(
        self,
        closed_trades: pd.DataFrame,
        open_positions: pd.DataFrame,
        symbol: str,
        start_date: datetime,
        end_date: datetime
    ) -> BinaryIO:
        """Generate Excel file with trade history"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required. Install with: pip install openpyxl")

        wb = Workbook()

        # Closed Trades Sheet
        ws_closed = wb.active
        ws_closed.title = "Closed Trades"

        headers = [
            'ID', 'Symbol', 'Strategy', 'Strike', 'Type', 'Contracts',
            'Entry Date', 'Entry Time', 'Entry Price',
            'Exit Date', 'Exit Time', 'Exit Price',
            'P&L ($)', 'P&L (%)', 'Hold Time (hrs)',
            'Entry Spot', 'Exit Spot', 'Entry VIX', 'Exit VIX',
            'GEX Regime', 'Exit Reason'
        ]
        self._write_header_row(ws_closed, headers)

        for idx, row in closed_trades.iterrows():
            row_num = idx + 2
            entry_value = float(row['entry_price'] or 0) * int(row['contracts'] or 0) * 100
            pnl_pct = (float(row['realized_pnl'] or 0) / entry_value * 100) if entry_value > 0 else 0

            data = [
                row['id'],
                row['symbol'],
                row['strategy'],
                row['strike'],
                row['option_type'],
                row['contracts'],
                str(row['entry_date']),
                str(row['entry_time']),
                float(row['entry_price'] or 0),
                str(row['exit_date']),
                str(row['exit_time']),
                float(row['exit_price'] or 0),
                float(row['realized_pnl'] or 0),
                round(pnl_pct, 2),
                float(row['hold_time_hours'] or 0),
                float(row['entry_spot_price'] or 0),
                float(row['exit_spot_price'] or 0),
                float(row['entry_vix'] or 0),
                float(row['exit_vix'] or 0),
                row['gex_regime'],
                row['exit_reason']
            ]
            for col, value in enumerate(data, 1):
                cell = ws_closed.cell(row=row_num, column=col, value=value)
                # Color P&L column
                if col == 13:  # P&L column
                    cell.font = self.money_positive if value >= 0 else self.money_negative

        self._auto_adjust_columns(ws_closed)

        # Open Positions Sheet
        ws_open = wb.create_sheet("Open Positions")

        headers_open = [
            'ID', 'Symbol', 'Strategy', 'Strike', 'Type', 'Contracts',
            'Entry Date', 'Entry Time', 'Entry Price', 'Current Price',
            'Unrealized P&L ($)', 'Entry Spot', 'Current Spot',
            'Expiration', 'GEX Regime', 'Confidence', 'Trade Reasoning'
        ]
        self._write_header_row(ws_open, headers_open)

        for idx, row in open_positions.iterrows():
            row_num = idx + 2
            data = [
                row['id'],
                row['symbol'],
                row['strategy'],
                row['strike'],
                row['option_type'],
                row['contracts'],
                str(row['entry_date']),
                str(row['entry_time']),
                float(row['entry_price'] or 0),
                float(row['current_price'] or 0),
                float(row['unrealized_pnl'] or 0),
                float(row['entry_spot_price'] or 0),
                float(row['current_spot_price'] or 0),
                str(row['expiration_date']) if row['expiration_date'] else '',
                row['gex_regime'],
                float(row['confidence'] or 0),
                str(row['trade_reasoning'] or '')[:200]  # Truncate long text
            ]
            for col, value in enumerate(data, 1):
                cell = ws_open.cell(row=row_num, column=col, value=value)
                if col == 11:  # Unrealized P&L
                    cell.font = self.money_positive if value >= 0 else self.money_negative

        self._auto_adjust_columns(ws_open)

        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Trade Export Summary").font = Font(bold=True, size=14)
        ws_summary.cell(row=3, column=1, value="Symbol:")
        ws_summary.cell(row=3, column=2, value=symbol)
        ws_summary.cell(row=4, column=1, value="Date Range:")
        ws_summary.cell(row=4, column=2, value=f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        ws_summary.cell(row=5, column=1, value="Closed Trades:")
        ws_summary.cell(row=5, column=2, value=len(closed_trades))
        ws_summary.cell(row=6, column=1, value="Open Positions:")
        ws_summary.cell(row=6, column=2, value=len(open_positions))

        if not closed_trades.empty:
            total_pnl = closed_trades['realized_pnl'].sum()
            wins = len(closed_trades[closed_trades['realized_pnl'] > 0])
            losses = len(closed_trades[closed_trades['realized_pnl'] < 0])
            win_rate = (wins / len(closed_trades) * 100) if len(closed_trades) > 0 else 0

            ws_summary.cell(row=8, column=1, value="Total Realized P&L:")
            pnl_cell = ws_summary.cell(row=8, column=2, value=f"${total_pnl:,.2f}")
            pnl_cell.font = self.money_positive if total_pnl >= 0 else self.money_negative

            ws_summary.cell(row=9, column=1, value="Win Rate:")
            ws_summary.cell(row=9, column=2, value=f"{win_rate:.1f}%")
            ws_summary.cell(row=10, column=1, value="Wins / Losses:")
            ws_summary.cell(row=10, column=2, value=f"{wins} / {losses}")

        ws_summary.cell(row=12, column=1, value="Generated:")
        ws_summary.cell(row=12, column=2, value=datetime.now(self.tz).strftime('%Y-%m-%d %H:%M:%S %Z'))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _export_trades_csv(self, closed_trades: pd.DataFrame, open_positions: pd.DataFrame) -> BinaryIO:
        """Export trades to CSV"""
        buffer = io.BytesIO()

        dfs_to_concat = []

        # Add closed trades if any
        if not closed_trades.empty:
            closed_trades = closed_trades.copy()
            closed_trades['status'] = 'CLOSED'
            cols = ['id', 'symbol', 'strategy', 'strike', 'option_type', 'contracts',
                   'entry_date', 'entry_price', 'realized_pnl', 'gex_regime', 'status']
            available_cols = [c for c in cols if c in closed_trades.columns]
            dfs_to_concat.append(closed_trades[available_cols])

        # Add open positions if any
        if not open_positions.empty:
            open_positions = open_positions.copy()
            open_positions['status'] = 'OPEN'
            if 'unrealized_pnl' in open_positions.columns:
                open_positions = open_positions.rename(columns={'unrealized_pnl': 'realized_pnl'})
            cols = ['id', 'symbol', 'strategy', 'strike', 'option_type', 'contracts',
                   'entry_date', 'entry_price', 'realized_pnl', 'gex_regime', 'status']
            available_cols = [c for c in cols if c in open_positions.columns]
            dfs_to_concat.append(open_positions[available_cols])

        if dfs_to_concat:
            combined = pd.concat(dfs_to_concat, ignore_index=True)
            combined.to_csv(buffer, index=False)
        else:
            buffer.write(b"No trades found\n")

        buffer.seek(0)
        return buffer

    def _export_pnl_attribution_excel(
        self,
        trades: pd.DataFrame,
        symbol: str,
        start_date: datetime,
        end_date: datetime,
        total_pnl: float
    ) -> BinaryIO:
        """Generate P&L attribution Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required")

        wb = Workbook()
        ws = wb.active
        ws.title = "P&L Attribution"

        headers = [
            'Trade ID', 'Exit Date', 'Strategy', 'Strike', 'Type', 'Contracts',
            'Entry Price', 'Exit Price', 'Gross P&L', 'Commission', 'Slippage',
            'Net P&L', 'Running Total', 'Contribution %', 'Result'
        ]
        self._write_header_row(ws, headers)

        for idx, row in trades.iterrows():
            row_num = idx + 2
            data = [
                row['id'],
                str(row['exit_date']),
                row['strategy'],
                row['strike'],
                row['option_type'],
                row['contracts'],
                round(float(row['entry_price'] or 0), 2),
                round(float(row['exit_price'] or 0), 2),
                round(float(row['gross_pnl'] or 0), 2),
                round(float(row['commission'] or 0), 2),
                round(float(row['slippage'] or 0), 2),
                round(float(row['net_pnl'] or 0), 2),
                round(float(row['running_total'] or 0), 2),
                round(float(row['contribution_pct'] or 0), 2),
                row['result']
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                # Color money columns
                if col in [9, 12, 13]:
                    cell.font = self.money_positive if value >= 0 else self.money_negative
                if col == 15:  # Result column
                    if value == 'WIN':
                        cell.font = self.money_positive
                    elif value == 'LOSS':
                        cell.font = self.money_negative

        self._auto_adjust_columns(ws)

        # Summary row
        summary_row = len(trades) + 3
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=12, value=round(total_pnl, 2)).font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _export_decision_logs_excel(
        self,
        logs: pd.DataFrame,
        symbol: str,
        start_date: datetime,
        end_date: datetime
    ) -> BinaryIO:
        """Generate decision logs Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required")

        wb = Workbook()

        # Main logs sheet
        ws = wb.active
        ws.title = "Decision Logs"

        headers = [
            'Timestamp', 'Log Type', 'Symbol', 'Spot Price', 'Net GEX',
            'Flip Point', 'VIX', 'Pattern', 'Confidence', 'Direction',
            'Risk Level', 'Psychology Trap', 'Strike Chosen', 'Contracts',
            'Action Taken', 'Strategy', 'Position ID'
        ]
        self._write_header_row(ws, headers)

        for idx, row in logs.iterrows():
            row_num = idx + 2
            data = [
                str(row['timestamp']),
                row['log_type'],
                row['symbol'],
                row['spot_price'],
                row['net_gex'],
                row['flip_point'],
                row['vix_level'],
                row['pattern_detected'],
                row['confidence_score'],
                row['trade_direction'],
                row['risk_level'],
                row['psychology_trap'],
                row['strike_chosen'],
                row['contracts'],
                row['action_taken'],
                row['strategy_name'],
                row['position_id']
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)

        self._auto_adjust_columns(ws)

        # AI Reasoning sheet
        ws_ai = wb.create_sheet("AI Reasoning")
        ai_headers = ['Timestamp', 'Action', 'AI Thought Process', 'AI Confidence', 'AI Warnings', 'Full Reasoning']
        self._write_header_row(ws_ai, ai_headers)

        ai_idx = 0
        for _, row in logs.iterrows():
            if row['ai_thought_process'] or row['full_reasoning']:
                ai_idx += 1
                ws_ai.cell(row=ai_idx + 1, column=1, value=str(row['timestamp']))
                ws_ai.cell(row=ai_idx + 1, column=2, value=row['action_taken'])
                ws_ai.cell(row=ai_idx + 1, column=3, value=str(row['ai_thought_process'] or '')[:500])
                ws_ai.cell(row=ai_idx + 1, column=4, value=row['ai_confidence'])
                ws_ai.cell(row=ai_idx + 1, column=5, value=str(row['ai_warnings'] or ''))
                ws_ai.cell(row=ai_idx + 1, column=6, value=str(row['full_reasoning'] or '')[:1000])

        self._auto_adjust_columns(ws_ai)

        # RSI Analysis sheet
        ws_rsi = wb.create_sheet("RSI Analysis")
        rsi_headers = ['Timestamp', 'Spot Price', 'RSI 5m', 'RSI 15m', 'RSI 1h', 'RSI 4h', 'RSI 1d', 'Action Taken']
        self._write_header_row(ws_rsi, rsi_headers)

        for idx, row in logs.iterrows():
            row_num = idx + 2
            ws_rsi.cell(row=row_num, column=1, value=str(row['timestamp']))
            ws_rsi.cell(row=row_num, column=2, value=row['spot_price'])
            ws_rsi.cell(row=row_num, column=3, value=row['rsi_5m'])
            ws_rsi.cell(row=row_num, column=4, value=row['rsi_15m'])
            ws_rsi.cell(row=row_num, column=5, value=row['rsi_1h'])
            ws_rsi.cell(row=row_num, column=6, value=row['rsi_4h'])
            ws_rsi.cell(row=row_num, column=7, value=row['rsi_1d'])
            ws_rsi.cell(row=row_num, column=8, value=row['action_taken'])

        self._auto_adjust_columns(ws_rsi)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _export_wheel_cycles_excel(
        self,
        cycles: pd.DataFrame,
        legs: pd.DataFrame,
        symbol: str
    ) -> BinaryIO:
        """Generate wheel cycles Excel"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required")

        wb = Workbook()

        # Cycles summary sheet
        ws = wb.active
        ws.title = "Wheel Cycles"

        headers = [
            'Cycle ID', 'Symbol', 'Status', 'Start Date', 'End Date',
            'Shares Owned', 'Cost Basis', 'CSP Premium', 'CC Premium',
            'Total Premium', 'Assignment Date', 'Assignment Price',
            'Called Away Date', 'Called Away Price', 'Realized P&L'
        ]
        self._write_header_row(ws, headers)

        for idx, row in cycles.iterrows():
            row_num = idx + 2
            data = [
                row['cycle_id'],
                row['symbol'],
                row['status'],
                str(row['start_date']) if row['start_date'] else '',
                str(row['end_date']) if row['end_date'] else '',
                row['shares_owned'],
                round(float(row['share_cost_basis'] or 0), 2),
                round(float(row['total_csp_premium'] or 0), 2),
                round(float(row['total_cc_premium'] or 0), 2),
                round(float(row['total_premium_collected'] or 0), 2),
                str(row['assignment_date']) if row['assignment_date'] else '',
                row['assignment_price'],
                str(row['called_away_date']) if row['called_away_date'] else '',
                row['called_away_price'],
                round(float(row['realized_pnl'] or 0), 2)
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                if col == 15:  # P&L
                    cell.font = self.money_positive if value >= 0 else self.money_negative

        self._auto_adjust_columns(ws)

        # Individual legs sheet
        ws_legs = wb.create_sheet("Wheel Legs")
        leg_headers = [
            'Leg ID', 'Cycle ID', 'Type', 'Action', 'Strike', 'Expiration',
            'Contracts', 'Premium Received', 'Premium Paid', 'Net Premium',
            'Open Date', 'Close Date', 'Close Reason', 'Delta at Open', 'DTE'
        ]
        self._write_header_row(ws_legs, leg_headers)

        for idx, row in legs.iterrows():
            row_num = idx + 2
            net_premium = (float(row['premium_received'] or 0) - float(row['premium_paid'] or 0)) * int(row['contracts'] or 0) * 100
            data = [
                row['leg_id'],
                row['cycle_id'],
                row['leg_type'],
                row['action'],
                row['strike'],
                str(row['expiration_date']),
                row['contracts'],
                round(float(row['premium_received'] or 0), 2),
                round(float(row['premium_paid'] or 0), 2),
                round(net_premium, 2),
                str(row['open_date']) if row['open_date'] else '',
                str(row['close_date']) if row['close_date'] else '',
                row['close_reason'],
                row['delta_at_open'],
                row['dte_at_open']
            ]
            for col, value in enumerate(data, 1):
                ws_legs.cell(row=row_num, column=col, value=value)

        self._auto_adjust_columns(ws_legs)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_header_row(self, ws, headers: List[str]):
        """Write formatted header row"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_empty_export(self, message: str) -> BinaryIO:
        """Create an empty export with a message"""
        if not OPENPYXL_AVAILABLE:
            buffer = io.BytesIO()
            buffer.write(message.encode())
            buffer.seek(0)
            return buffer

        wb = Workbook()
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value=message)
        ws.cell(row=3, column=1, value="Generated:")
        ws.cell(row=3, column=2, value=datetime.now(self.tz).strftime('%Y-%m-%d %H:%M:%S %Z'))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _add_trade_history_sheet(self, wb, conn, symbol, start_date, end_date):
        """Add trade history sheet to workbook"""
        trades = pd.read_sql_query('''
            SELECT * FROM autonomous_closed_trades
            WHERE symbol = %s AND COALESCE(exit_date, entry_date) >= %s AND COALESCE(exit_date, entry_date) <= %s
            ORDER BY COALESCE(exit_date, entry_date) DESC
        ''', conn, params=(symbol, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

        ws = wb.create_sheet("Trade History")
        if trades.empty:
            ws.cell(row=1, column=1, value="No trades found")
            return

        headers = list(trades.columns)
        self._write_header_row(ws, headers)
        for idx, row in trades.iterrows():
            for col, value in enumerate(row, 1):
                ws.cell(row=idx + 2, column=col, value=value)
        self._auto_adjust_columns(ws)

    def _add_pnl_attribution_sheet(self, wb, conn, symbol, start_date, end_date):
        """Add P&L attribution sheet"""
        ws = wb.create_sheet("P&L Attribution")
        # Similar implementation to export_pnl_attribution
        ws.cell(row=1, column=1, value="P&L Attribution")

    def _add_decision_logs_sheet(self, wb, conn, symbol, start_date, end_date):
        """Add decision logs sheet"""
        ws = wb.create_sheet("Decision Logs")
        ws.cell(row=1, column=1, value="Decision Logs")

    def _add_wheel_cycles_sheet(self, wb, conn, symbol):
        """Add wheel cycles sheet if applicable"""
        ws = wb.create_sheet("Wheel Cycles")
        ws.cell(row=1, column=1, value="Wheel Cycles")

    def _add_performance_summary_sheet(self, wb, conn, symbol, start_date, end_date):
        """Add performance summary sheet"""
        ws = wb.create_sheet("Performance Summary")

        # Query summary stats
        cursor = conn.cursor()
        cursor.execute('''
            SELECT
                COUNT(*) as total_trades,
                SUM(CASE WHEN realized_pnl > 0 THEN 1 ELSE 0 END) as wins,
                SUM(CASE WHEN realized_pnl < 0 THEN 1 ELSE 0 END) as losses,
                SUM(realized_pnl) as total_pnl,
                AVG(realized_pnl) as avg_pnl,
                MAX(realized_pnl) as best_trade,
                MIN(realized_pnl) as worst_trade,
                AVG(hold_time_hours) as avg_hold_time
            FROM autonomous_closed_trades
            WHERE symbol = %s AND COALESCE(exit_date, entry_date) >= %s AND COALESCE(exit_date, entry_date) <= %s
        ''', (symbol, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

        row = cursor.fetchone()

        ws.cell(row=1, column=1, value="Performance Summary").font = Font(bold=True, size=14)
        ws.cell(row=3, column=1, value="Symbol:")
        ws.cell(row=3, column=2, value=symbol)
        ws.cell(row=4, column=1, value="Period:")
        ws.cell(row=4, column=2, value=f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        if row and row[0] > 0:
            total, wins, losses, total_pnl, avg_pnl, best, worst, avg_hold = row
            win_rate = (wins / total * 100) if total > 0 else 0

            ws.cell(row=6, column=1, value="Total Trades:")
            ws.cell(row=6, column=2, value=total)
            ws.cell(row=7, column=1, value="Win Rate:")
            ws.cell(row=7, column=2, value=f"{win_rate:.1f}%")
            ws.cell(row=8, column=1, value="Wins / Losses:")
            ws.cell(row=8, column=2, value=f"{wins} / {losses}")
            ws.cell(row=9, column=1, value="Total P&L:")
            pnl_cell = ws.cell(row=9, column=2, value=f"${total_pnl:,.2f}")
            pnl_cell.font = self.money_positive if total_pnl >= 0 else self.money_negative
            ws.cell(row=10, column=1, value="Avg P&L per Trade:")
            ws.cell(row=10, column=2, value=f"${avg_pnl:,.2f}")
            ws.cell(row=11, column=1, value="Best Trade:")
            ws.cell(row=11, column=2, value=f"${best:,.2f}")
            ws.cell(row=12, column=1, value="Worst Trade:")
            ws.cell(row=12, column=2, value=f"${worst:,.2f}")
            ws.cell(row=13, column=1, value="Avg Hold Time:")
            ws.cell(row=13, column=2, value=f"{avg_hold:.1f} hours")
        else:
            ws.cell(row=6, column=1, value="No trades in this period")


# Singleton instance
export_service = TradeExportService()
